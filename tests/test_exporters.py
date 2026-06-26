import json

from openpyxl import load_workbook

from purchase_prep_assistant.exporters import load_plan, write_all_exports


def test_write_all_exports(tmp_path) -> None:
    plan = load_plan("sample_data/purchase_plan.json")
    files = write_all_exports(plan, tmp_path)
    names = {path.name for path in files}
    assert "prime-sale-business-api-sample.csv" in names
    assert "prime-sale-business-api-sample.xlsx" in names
    assert "prime-sale-business-api-sample-checklist.txt" in names
    assert "prime-sale-business-api-sample-amazon-business-template.json" in names

    workbook = load_workbook(tmp_path / "prime-sale-business-api-sample.xlsx")
    assert "Items" in workbook.sheetnames
    assert "API Setup" in workbook.sheetnames

    template = json.loads(
        (tmp_path / "prime-sale-business-api-sample-amazon-business-template.json").read_text(
            encoding="utf-8"
        )
    )
    assert template["cart_add_items"]["payload"]["items"][0]["quantity"] == 2
    assert "customer-a" in template["ordering_api_trial_payloads"]
