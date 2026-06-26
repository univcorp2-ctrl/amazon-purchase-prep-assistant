"""Export purchase automation plans to review-friendly and API-ready files."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from purchase_prep_assistant.business_payloads import build_cart_add_items_payload, build_order_payload
from purchase_prep_assistant.models import PurchasePlan
from purchase_prep_assistant.safety import SAFETY_POLICY_TEXT, assert_safe_plan, inspect_plan


def load_plan(path: str | Path) -> PurchasePlan:
    """Load a purchase plan from JSON."""
    with Path(path).open("r", encoding="utf-8") as file:
        data = json.load(file)
    return PurchasePlan.model_validate(data)


def write_csv(plan: PurchasePlan, output_path: str | Path) -> Path:
    """Write item-level purchase preparation data to CSV."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "project_name",
                "workflow_mode",
                "business_region",
                "product_name",
                "url",
                "asin",
                "buying_option_identifier",
                "quantity",
                "max_unit_price_jpy",
                "note",
            ],
        )
        writer.writeheader()
        for product in plan.products:
            writer.writerow(
                {
                    "project_name": plan.project_name,
                    "workflow_mode": plan.workflow_mode,
                    "business_region": plan.business_region,
                    "product_name": product.name,
                    "url": product.url,
                    "asin": product.asin or "",
                    "buying_option_identifier": product.buying_option_identifier or "",
                    "quantity": product.quantity,
                    "max_unit_price_jpy": product.max_unit_price_jpy or "",
                    "note": product.note,
                }
            )
    return output


def _add_header(ws: Any, values: list[str]) -> None:
    ws.append(values)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")


def write_excel(plan: PurchasePlan, output_path: str | Path) -> Path:
    """Write an Excel workbook with products, recipients, allocations, and API setup notes."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    report = inspect_plan(plan)
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Summary"
    _add_header(summary, ["Field", "Value"])
    summary.append(["Project", plan.project_name])
    summary.append(["Workflow mode", plan.workflow_mode])
    summary.append(["Business region", plan.business_region])
    summary.append(["Currency", plan.currency_code])
    summary.append(["Plan OK", "yes" if report.ok else "no"])
    summary.append(["Warnings", " | ".join(report.warnings) if report.warnings else "none"])

    items = workbook.create_sheet("Items")
    _add_header(
        items,
        [
            "Product",
            "URL",
            "ASIN",
            "Buying option identifier",
            "Quantity",
            "Max unit price JPY",
            "Note",
        ],
    )
    for product in plan.products:
        items.append(
            [
                product.name,
                product.url,
                product.asin or "",
                product.buying_option_identifier or "",
                product.quantity,
                product.max_unit_price_jpy or "",
                product.note,
            ]
        )

    recipients = workbook.create_sheet("Recipients")
    _add_header(
        recipients,
        [
            "Label",
            "Name",
            "Postal code",
            "Prefecture",
            "City",
            "Line1",
            "Line2",
            "Line3",
            "Phone",
            "Company",
            "Country",
            "Buyer email",
            "External address ID",
            "Note",
        ],
    )
    for recipient in plan.recipients:
        recipients.append(
            [
                recipient.label,
                recipient.recipient_name,
                recipient.postal_code,
                recipient.prefecture,
                recipient.city,
                recipient.line1,
                recipient.line2,
                recipient.line3,
                recipient.phone,
                recipient.company_name,
                recipient.country_code,
                recipient.buyer_email,
                recipient.external_address_id,
                recipient.note,
            ]
        )

    allocations = workbook.create_sheet("Allocations")
    _add_header(allocations, ["Product", "Recipient", "Quantity"])
    for allocation in plan.allocations:
        allocations.append([allocation.product_name, allocation.recipient_label, allocation.quantity])

    setup = workbook.create_sheet("API Setup")
    _add_header(setup, ["Key", "Value"])
    setup.append(["Project scope", SAFETY_POLICY_TEXT])
    setup.append(["Buying group reference", plan.buying_group_reference])
    setup.append(["Payment method reference", plan.payment_method_reference])
    setup.append(["Buyer reference", plan.buyer_reference])
    setup.append(["Purchase order number", plan.purchase_order_number])

    for worksheet in workbook.worksheets:
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 80)

    workbook.save(output)
    return output


def write_checklist(plan: PurchasePlan, output_path: str | Path) -> Path:
    """Write a human-readable operations checklist."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    report = inspect_plan(plan)

    lines = [
        f"# Operations checklist: {plan.project_name}",
        "",
        "## Project / プロジェクト",
        f"- Workflow mode: {plan.workflow_mode}",
        f"- Region: {plan.business_region}",
        f"- Currency: {plan.currency_code}",
        "",
        "## Products / 商品",
    ]
    for index, product in enumerate(plan.products, start=1):
        lines.extend(
            [
                f"{index}. {product.name}",
                f"   - URL: {product.url}",
                f"   - ASIN: {product.asin or 'ASIN_REQUIRED'}",
                f"   - Buying option: {product.buying_option_identifier or 'BUYING_OPTION_IDENTIFIER_REQUIRED'}",
                f"   - Quantity: {product.quantity}",
                f"   - Max unit price JPY: {product.max_unit_price_jpy or 'not set'}",
            ]
        )

    lines.extend(["", "## Recipients / 顧客・配送先"])
    for recipient in plan.recipients:
        lines.append(
            f"- {recipient.label}: {recipient.recipient_name}, {recipient.postal_code}, {recipient.prefecture}{recipient.city}{recipient.line1}"
        )

    lines.extend(["", "## Allocations / 割当"])
    if plan.allocations:
        for allocation in plan.allocations:
            lines.append(
                f"- {allocation.product_name}: {allocation.quantity} item(s) -> {allocation.recipient_label}"
            )
    else:
        lines.append("- All products are applied to the selected recipient when building an order payload.")

    lines.extend(["", "## Warnings / 確認事項"])
    lines.extend([f"- {warning}" for warning in report.warnings] or ["- none"])

    output.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return output


def build_amazon_business_cart_template(plan: PurchasePlan) -> dict[str, Any]:
    """Build a Cart API addItems template."""
    return {
        "api": "Amazon Business Cart API",
        "path": "/cart/2025-04-30/carts/{cartId}/items",
        "method": "POST",
        "region": plan.business_region,
        "payload": build_cart_add_items_payload(plan),
    }


def build_amazon_business_order_templates(plan: PurchasePlan) -> dict[str, Any]:
    """Build one Ordering API payload per recipient."""
    return {
        recipient.label: build_order_payload(plan, recipient.label, trial=True)
        for recipient in plan.recipients
    }


def write_business_api_template(plan: PurchasePlan, output_path: str | Path) -> Path:
    """Write a JSON bundle for official Amazon Business API integration."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "cart_add_items": build_amazon_business_cart_template(plan),
        "ordering_api_trial_payloads": build_amazon_business_order_templates(plan),
    }
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return output


def write_all_exports(plan: PurchasePlan, output_dir: str | Path) -> list[Path]:
    """Validate a plan and write all supported outputs."""
    assert_safe_plan(plan)
    directory = Path(output_dir)
    directory.mkdir(parents=True, exist_ok=True)
    safe_name = plan.project_name.replace(" ", "-").lower()
    return [
        write_csv(plan, directory / f"{safe_name}.csv"),
        write_excel(plan, directory / f"{safe_name}.xlsx"),
        write_checklist(plan, directory / f"{safe_name}-checklist.txt"),
        write_business_api_template(plan, directory / f"{safe_name}-amazon-business-template.json"),
    ]
